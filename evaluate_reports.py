"""
Evaluate radiology reports: compare ground truth from test_metadata.csv
against predicted reports using RadEval metrics and optionally CRIMSON.

Predicted reports can be:
  - A column from the CSV (e.g. a model's output)
  - A fixed "normal report" from a text file (applied to every row)

Usage examples:
  # Normal-report baseline, first 50, with CRIMSON
  python evaluate_reports.py --n 50 --normal-report normal_report.txt --crimson

  # Normal-report baseline, all rows, RadEval only
  python evaluate_reports.py --normal-report normal_report.txt

  # Predicted column from CSV
  python evaluate_reports.py --pred-column Predicted --n 100

  # Custom output path
  python evaluate_reports.py --normal-report normal_report.txt --output results.xlsx
"""

import argparse
import json
import sys
import time
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor

# Ensure this project directory is on sys.path so local packages like `RadEval/` import correctly
BASE_DIR = Path(__file__).resolve().parent
if str(BASE_DIR) not in sys.path:
    sys.path.insert(0, str(BASE_DIR))

# Add RadEval directory to sys.path for its internal imports
RADEVAL_DIR = BASE_DIR / "RadEval"
if RADEVAL_DIR.exists() and str(RADEVAL_DIR) not in sys.path:
    sys.path.insert(0, str(RADEVAL_DIR))

import pandas as pd
import numpy as np
from tqdm import tqdm
from RadEval.RadEval import RadEval


def flatten_radeval_scores(scores):
    """Flatten RadEval output dict into a single-level {metric: value} dict."""
    flat = {}
    for key, value in scores.items():
        if key.startswith('rouge') and key != 'rougeL':
            continue
        if key.startswith('chexbert') and key != 'chexbert-5_micro avg_f1-score':
            continue
        if key.startswith('radgraph') and key not in ['radgraph_complete']:
            continue

        if isinstance(value, dict):
            if 'bleu' in key.lower():
                flat[key] = value.get('bleu-4', list(value.values())[0])
            elif 'rouge' in key.lower():
                flat['rougeL'] = value.get('rougeL', value.get('rouge-l', list(value.values())[0]))
            elif 'chexbert' in key.lower():
                flat['chexbert'] = value.get('chexbert-5_micro avg_f1-score', list(value.values())[0])
            else:
                flat[key] = list(value.values())[0]
        else:
            flat[key] = value
    return flat


def main():
    parser = argparse.ArgumentParser(
        description="Evaluate predicted radiology reports against ground truth."
    )
    parser.add_argument(
        "--input", default="data/rexgradient/test_metadata.csv",
        help="Path to ground truth CSV (default: data/rexgradient/test_metadata.csv)"
    )
    parser.add_argument(
        "--gt-column", default="Findings",
        help="Column name for ground truth reports (default: Findings)"
    )
    parser.add_argument(
        "--pred-column", default=None,
        help="Column name for predicted reports (mutually exclusive with --normal-report)"
    )
    parser.add_argument(
        "--normal-report", default=None,
        help="Path to a text file whose content is used as the predicted report for every row"
    )
    parser.add_argument(
        "--n", type=int, default=None,
        help="Only evaluate the first N rows (default: all)"
    )
    parser.add_argument(
        "--crimson", action="store_true",
        help="Also compute CRIMSON scores (requires generate_score.py)"
    )
    parser.add_argument(
        "--no-radeval", action="store_true",
        help="Skip RadEval metrics (useful if you only want CRIMSON)"
    )
    parser.add_argument(
        "--cache_dir", default=None,
        help="Cache directory for RadEval model downloads"
    )
    parser.add_argument(
        "--output", default=None,
        help="Output path (.xlsx or .csv). Default: data/evaluation_results_<timestamp>.xlsx"
    )
    parser.add_argument(
        "--crimson-details", action="store_true",
        help="Also output total/weighted error counts from CRIMSON (implies --crimson)"
    )
    parser.add_argument(
        "--max-workers", type=int, default=8,
        help="Max concurrent CRIMSON API calls (default: 8)"
    )
    args = parser.parse_args()

    # --crimson-details implies --crimson
    if args.crimson_details:
        args.crimson = True

    # Validate arguments
    if args.pred_column and args.normal_report:
        parser.error("Use either --pred-column or --normal-report, not both.")
    if not args.pred_column and not args.normal_report:
        parser.error("Must specify either --pred-column or --normal-report.")

    df = pd.read_csv(args.input)
    if args.n:
        df = df.head(args.n)
    print(f"Loaded {len(df)} rows from {args.input}")

    gt_reports = df[args.gt_column].tolist()

    if args.normal_report:
        normal_text = Path(args.normal_report).read_text().strip()
        pred_reports = [normal_text] * len(df)
        pred_source = f"normal_report ({args.normal_report})"
    else:
        pred_reports = df[args.pred_column].tolist()
        pred_source = f"column '{args.pred_column}'"

    print(f"Ground truth: column '{args.gt_column}' | Predicted: {pred_source}")
    print(f"Evaluating {len(gt_reports)} pairs...")

    results = []
    for i in range(len(gt_reports)):
        results.append({
            "id": df.iloc[i].get("id", i),
            "ground_truth": gt_reports[i],
            "predicted": pred_reports[i],
        })

    radeval_metric_names = []
    if not args.no_radeval:
        print("\nComputing RadEval metrics...")
        evaluator_kwargs = {
            'do_radgraph': True,
            'do_bleu': True,
            'do_bertscore': True,
            'do_green': True,
            'do_chexbert': True,
            'do_ratescore': True,
            'do_radcliq': True,
            'do_rouge': True,
        }
        if args.cache_dir:
            evaluator_kwargs['cache_dir'] = args.cache_dir

        radeval_scorer = RadEval(**evaluator_kwargs)

        t0 = time.time()
        green_error_cols = {
            "(a) False report of a finding in the candidate": "green_false_finding",
            "(b) Missing a finding present in the reference": "green_missing_finding",
            "(c) Misidentification of a finding's anatomic location/position": "green_location_error",
            "(d) Misassessment of the severity of a finding": "green_severity_error",
            "(e) Mentioning a comparison that isn't in the reference": "green_false_comparison",
            "(f) Omitting a comparison detailing a change from a prior study": "green_missing_comparison",
            "Matched Findings": "green_matched_findings",
        }
        for i, row in enumerate(tqdm(results, desc="RadEval")):
            scores = radeval_scorer(refs=[row["ground_truth"]], hyps=[row["predicted"]])
            flat = flatten_radeval_scores(scores)
            row.update(flat)

            if hasattr(radeval_scorer, 'green_results_df'):
                gdf = radeval_scorer.green_results_df
                for orig_col, new_col in green_error_cols.items():
                    row[new_col] = int(gdf.iloc[0].get(orig_col, 0))
                row["green_total_sig_errors"] = sum(
                    int(gdf.iloc[0].get(orig_col, 0))
                    for orig_col in list(green_error_cols.keys())[:-1]
                )

            if i == 0:
                radeval_metric_names = list(flat.keys())

        print(f"  RadEval done in {time.time() - t0:.1f}s")

    if args.crimson:
        from CRIMSON.generate_score import CRIMSONScore
        scorer = CRIMSONScore()

        print(f"\nComputing CRIMSON scores (max_workers={args.max_workers})...")
        t0 = time.time()

        def score_one(idx):
            row = results[idx]
            for attempt in range(3):
                try:
                    result = scorer.evaluate(row["ground_truth"], row["predicted"])
                    return idx, result
                except Exception as e:
                    if attempt < 2:
                        time.sleep(2 ** attempt)
                        continue
                    print(f"  CRIMSON error on row {idx} (after 3 attempts): {e}")
                    return idx, None

        n_success, n_fail = 0, 0
        with ThreadPoolExecutor(max_workers=args.max_workers) as executor:
            futures = [executor.submit(score_one, i) for i in range(len(results))]
            for future in tqdm(futures, desc="CRIMSON"):
                idx, result = future.result()
                if result is None:
                    results[idx]["crimson_score"] = None
                    results[idx]["crimson_json"] = None
                    n_fail += 1
                    continue
                n_success += 1
                results[idx]["crimson_score"] = result["crimson_score"]
                results[idx]["crimson_json"] = json.dumps(result["raw_evaluation"], indent=4)

                if args.crimson_details:
                    ec = result.get("error_counts", {})
                    results[idx]["crimson_total_errors"] = (
                        ec.get("false_findings", 0)
                        + ec.get("missing_findings", 0)
                        + ec.get("attribute_errors", 0)
                    )
                    wc = result.get("weighted_error_counts", {})
                    results[idx]["crimson_weighted_errors"] = (
                        wc.get("false_findings", 0)
                        + wc.get("missing_findings", 0)
                        + wc.get("attribute_errors", 0)
                    )
                    results[idx]["crimson_false_findings"] = ec.get("false_findings", 0)
                    results[idx]["crimson_missing_findings"] = ec.get("missing_findings", 0)
                    results[idx]["crimson_attribute_errors"] = ec.get("attribute_errors", 0)

        print(f"  CRIMSON done in {time.time() - t0:.1f}s ({n_success} succeeded, {n_fail} failed)")
        if n_fail > 0:
            print(f"  WARNING: {n_fail} rows failed — average will exclude them!")

    out_df = pd.DataFrame(results)

    print("\n" + "=" * 60)
    print("AVERAGE SCORES")
    print("=" * 60)

    avg_row = {"id": "AVERAGE", "ground_truth": "", "predicted": ""}
    numeric_cols = []

    for col in radeval_metric_names:
        if col in out_df.columns:
            vals = pd.to_numeric(out_df[col], errors="coerce")
            avg_val = vals.mean()
            avg_row[col] = round(avg_val, 4)
            numeric_cols.append(col)
            print(f"  {col}: {avg_val:.4f}")

    if args.crimson and "crimson_score" in out_df.columns:
        vals = pd.to_numeric(out_df["crimson_score"], errors="coerce")
        avg_val = vals.mean()
        avg_row["crimson_score"] = round(avg_val, 4)
        numeric_cols.append("crimson_score")
        print(f"  crimson_score: {avg_val:.4f}")

    if "green_total_sig_errors" in out_df.columns:
        green_avg_cols = [
            "green_false_finding", "green_missing_finding",
            "green_location_error", "green_severity_error",
            "green_false_comparison", "green_missing_comparison",
            "green_matched_findings", "green_total_sig_errors",
        ]
        for col in green_avg_cols:
            if col in out_df.columns:
                vals = pd.to_numeric(out_df[col], errors="coerce")
                avg_val = vals.mean()
                avg_row[col] = round(avg_val, 4)
                numeric_cols.append(col)
                print(f"  {col}: {avg_val:.4f}")

    print("=" * 60)

    avg_df = pd.DataFrame([avg_row])
    out_df = pd.concat([out_df, avg_df], ignore_index=True)

    if args.output:
        out_path = Path(args.output)
        
    out_path.parent.mkdir(parents=True, exist_ok=True)

    if out_path.suffix == ".xlsx":
        import openpyxl
        from openpyxl.utils import get_column_letter

        out_df.to_excel(out_path, index=False, engine="openpyxl")

        wb = openpyxl.load_workbook(out_path)
        ws = wb.active
        wrap_fmt = openpyxl.styles.Alignment(wrap_text=True, vertical="top")

        for col_idx, col_name in enumerate(out_df.columns, start=1):
            max_width = len(str(col_name))
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                cell = row[0]
                cell.alignment = wrap_fmt
                if cell.value is not None:
                    lines = str(cell.value).split("\n")
                    max_line = max(len(line) for line in lines)
                    max_width = max(max_width, min(max_line, 80))
            ws.column_dimensions[get_column_letter(col_idx)].width = max_width + 2

        wb.save(out_path)
    else:
        out_df.to_csv(out_path, index=False)

    print(f"\nResults saved to: {out_path}")
    print(f"Total rows: {len(out_df) - 1} + 1 average row")


if __name__ == "__main__":
    main()